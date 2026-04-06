import argparse
import json
import shutil
import os
from openpyxl import load_workbook

def generate_excel(input_json, template_path, output_path, month_name):
    """
    input_json: path to json file with structured data
    template_path: original XLSX template
    output_path: where to save the generated XLSX
    month_name: e.g. 'JAN/2026' to search for worksheet
    """
    with open(input_json, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Copy template to output first to keep styles
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)

    # Try to find the sheet
    if month_name not in wb.sheetnames:
        print(f"Sheet '{month_name}' not found. Using the first sheet.")
        ws = wb.active
    else:
        ws = wb[month_name]

    # LOGIC: Loop through rows to find headers and insert data
    # This is a simplified version. The AI using the skill will map categories to row ranges.
    # For a deterministic script, we can look for keywords like 'MENSALIDADES', 'SAÍDAS', etc.
    
    # Example insertion logic for 'GASTOS DIVERSOS'
    # For each entry in data['saidas'], find 'GASTOS DIVERSOS' header and append below it
    
    # TO BE REFINED BASED ON EXACT SHEET STRUCTURE OBSERVED
    # But for now, we'll implement a robust 'find and write' mechanism
    
    wb.save(output_path)
    print(f"Excel generated at: {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--template", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--month", required=True)
    args = parser.parse_args()
    
    generate_excel(args.input, args.template, args.output, args.month)
